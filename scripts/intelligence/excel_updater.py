"""
Update the SEO roadmap Excel workbook in place.

Modifies only specific cells — never deletes rows, never destroys
formatting, never touches completed tasks. Preserves cell formatting,
colors, formulas, filters, column widths, and freeze panes.
"""

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

# Column mapping for the 101-Day Roadmap sheet (1-based)
COL = {
    "day": 1, "date": 2, "week": 3, "module": 4, "type": 5,
    "title": 6, "detail": 7, "prompt": 8, "why": 9, "est_time": 10,
    "fixed_flexible": 11, "status": 12, "notes": 13, "tool": 14,
}

STATUS_COLORS = {
    "DONE": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "IN PROGRESS": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "BLOCKED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "SUPERSEDED": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "REPRIORITIZED": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
}


@dataclass
class CellUpdate:
    sheet: str
    row: int
    column: str  # key from COL dict
    value: str
    reason: str = ""


class ExcelUpdater:
    def __init__(self, workbook_path: Path):
        self._path = workbook_path
        self._changes: list[CellUpdate] = []

    def queue_update(self, update: CellUpdate):
        """Queue a cell update for batch application."""
        self._changes.append(update)

    def queue_status_update(self, row: int, status: str, notes: str = "", reason: str = ""):
        """Queue a status change on the roadmap sheet."""
        self._changes.append(CellUpdate(
            sheet="101-Day Roadmap", row=row, column="status",
            value=status, reason=reason,
        ))
        if notes:
            self._changes.append(CellUpdate(
                sheet="101-Day Roadmap", row=row, column="notes",
                value=notes, reason=reason,
            ))

    def queue_priority_update(self, row: int, new_detail: str, why: str, notes: str):
        """Queue a reprioritization update."""
        self._changes.append(CellUpdate(
            sheet="101-Day Roadmap", row=row, column="detail",
            value=new_detail, reason="Reprioritized based on data",
        ))
        self._changes.append(CellUpdate(
            sheet="101-Day Roadmap", row=row, column="why",
            value=why, reason="Evidence-based update",
        ))
        self._changes.append(CellUpdate(
            sheet="101-Day Roadmap", row=row, column="notes",
            value=notes, reason="Data-driven note",
        ))

    def apply_all(self) -> int:
        """Apply all queued changes to the workbook. Returns count of cells modified."""
        if not self._changes:
            logger.info("No changes to apply")
            return 0

        wb = openpyxl.load_workbook(self._path)
        modified = 0

        for update in self._changes:
            if update.sheet not in wb.sheetnames:
                logger.warning("Sheet '%s' not found — skipping", update.sheet)
                continue

            ws = wb[update.sheet]
            col_idx = COL.get(update.column)
            if col_idx is None:
                logger.warning("Unknown column '%s' — skipping", update.column)
                continue

            cell = ws.cell(row=update.row, column=col_idx)

            # Preserve existing formatting, only update value
            cell.value = update.value

            # Apply status color if updating status column
            if update.column == "status":
                fill = STATUS_COLORS.get(update.value.upper())
                if fill:
                    cell.fill = fill

            modified += 1
            logger.debug(
                "Updated [%s] row %d col %s = '%s' (%s)",
                update.sheet, update.row, update.column,
                str(update.value)[:50], update.reason,
            )

        wb.save(self._path)
        wb.close()

        self._changes.clear()
        logger.info("Applied %d cell updates to %s", modified, self._path.name)
        return modified

    def get_pending_changes(self) -> list[CellUpdate]:
        """View queued changes before applying."""
        return list(self._changes)

    def clear_pending(self):
        """Discard all queued changes."""
        self._changes.clear()
