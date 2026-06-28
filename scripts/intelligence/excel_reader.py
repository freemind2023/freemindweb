"""
Read the SEO roadmap Excel workbook preserving full structure.

Parses all three sheets (101-Day Roadmap, Audit Findings, Weekly Themes)
into structured Python objects for analysis and comparison.
"""

import logging
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

ROADMAP_COLUMNS = [
    "day", "date", "week", "module", "type", "title", "detail",
    "prompt", "why", "est_time", "fixed_flexible", "status",
    "notes", "tool",
]


@dataclass
class RoadmapTask:
    row_number: int  # Excel row (1-based, header=1, data starts at 2)
    day: int
    date: str
    week: int
    module: str
    task_type: str
    title: str
    detail: str
    prompt: str
    why: str
    est_time: str
    fixed_flexible: str
    status: str
    notes: str
    tool: str

    def is_done(self) -> bool:
        return str(self.status).strip().upper() == "DONE"

    def is_pending(self) -> bool:
        return not self.status or str(self.status).strip() == ""

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class AuditFinding:
    row_number: int
    category: str
    page: str
    finding: str
    status: str
    priority: str
    notes: str

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class WeeklyTheme:
    row_number: int
    week: int
    days: str
    phase: str
    focus: str
    deliverables: str
    mix: str

    def to_dict(self) -> dict:
        return asdict(self)


class ExcelReader:
    def __init__(self, workbook_path: Path):
        self._path = workbook_path

    def read_all(self) -> dict:
        """Read the entire workbook into structured objects."""
        wb = openpyxl.load_workbook(self._path)

        result = {
            "tasks": self._read_roadmap(wb["101-Day Roadmap"]),
            "audit_findings": self._read_audit(wb["Audit Findings"]),
            "weekly_themes": self._read_themes(wb["Weekly Themes"]),
            "sheet_names": wb.sheetnames,
        }

        wb.close()
        logger.info(
            "Read workbook: %d tasks, %d audit findings, %d weekly themes",
            len(result["tasks"]),
            len(result["audit_findings"]),
            len(result["weekly_themes"]),
        )
        return result

    def read_tasks(self) -> list[RoadmapTask]:
        """Read only the roadmap tasks."""
        wb = openpyxl.load_workbook(self._path)
        tasks = self._read_roadmap(wb["101-Day Roadmap"])
        wb.close()
        return tasks

    def get_done_tasks(self) -> list[RoadmapTask]:
        return [t for t in self.read_tasks() if t.is_done()]

    def get_pending_tasks(self) -> list[RoadmapTask]:
        return [t for t in self.read_tasks() if t.is_pending()]

    def get_next_30_days(self) -> list[RoadmapTask]:
        """Get the next 30 pending tasks (the actionable window)."""
        pending = self.get_pending_tasks()
        return pending[:30]

    def _read_roadmap(self, ws) -> list[RoadmapTask]:
        tasks = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            vals = [str(v) if v is not None else "" for v in row]
            while len(vals) < 14:
                vals.append("")
            tasks.append(RoadmapTask(
                row_number=row_idx,
                day=int(float(vals[0])) if vals[0] else 0,
                date=vals[1],
                week=int(float(vals[2])) if vals[2] else 0,
                module=vals[3],
                task_type=vals[4],
                title=vals[5],
                detail=vals[6],
                prompt=vals[7],
                why=vals[8],
                est_time=vals[9],
                fixed_flexible=vals[10],
                status=vals[11],
                notes=vals[12],
                tool=vals[13],
            ))
        return tasks

    def _read_audit(self, ws) -> list[AuditFinding]:
        findings = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if not row or all(v is None for v in row):
                continue
            vals = [str(v) if v is not None else "" for v in row]
            while len(vals) < 6:
                vals.append("")
            findings.append(AuditFinding(
                row_number=row_idx,
                category=vals[0],
                page=vals[1],
                finding=vals[2],
                status=vals[3],
                priority=vals[4],
                notes=vals[5],
            ))
        return findings

    def _read_themes(self, ws) -> list[WeeklyTheme]:
        themes = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            vals = [str(v) if v is not None else "" for v in row]
            while len(vals) < 6:
                vals.append("")
            themes.append(WeeklyTheme(
                row_number=row_idx,
                week=int(float(vals[0])) if vals[0] else 0,
                days=vals[1],
                phase=vals[2],
                focus=vals[3],
                deliverables=vals[4],
                mix=vals[5],
            ))
        return themes
