"""
Compress the 126-day SEO roadmap into 15 days.

Days 1-7 (DONE) remain untouched.
Days 8-15 are rebuilt as 8 mega-days, each covering
multiple original modules — prioritized by live GA4/GSC data.
"""

import json
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import DATA_DIR, BACKUPS_DIR

WORKBOOK = Path(__file__).parent.parent / "01_seo_audit_roadmap.xlsx"

# ── Evidence strings pulled from live data ──────────────────────────
# (from combined_pages.csv + queries.csv + intelligence report)

EVIDENCE = {
    "indexing": (
        "GSC: Only 15 of 31 tracked pages have search impressions. "
        "46 HTML files on site but many unindexed. "
        "GA4: 52 organic sessions/month — massive upside if indexed."
    ),
    "titles_meta": (
        "HTML scan: 43 of 46 titles exceed 60 chars (truncated in SERPs). "
        "8 pages have CTR < 2% with 10+ impressions. "
        "Top CTR gap: homepage at position 6.4 gets 1.4% CTR vs expected 5%."
    ),
    "content": (
        "161 technical issues detected across 46 pages. "
        "Multiple pages with missing H1, thin content (<800 words), "
        "missing FAQ schema. Blog: 78 sessions, 50% avg bounce rate."
    ),
    "keywords": (
        "82 unique queries in GSC. Avg position 35.0. "
        "'isbn registration india' at position 10 with 35 impressions — striking distance. "
        "'how to publish a book in india' — 81 impressions, position 72, 0 clicks. "
        "Brand confusion: 'freemind chatbot platform' from US = 200+ impressions, 0 clicks."
    ),
    "internal_links": (
        "16 pages with search impressions but no GA4 sessions (orphan/weak linking). "
        "3 pages indexed but getting 0 organic sessions. "
        "12 pages are 'direct_only' — no search visibility."
    ),
    "technical": (
        "Missing schema on multiple pages. Missing alt text on images. "
        "Missing canonical tags detected. Multiple H1 issues. "
        "PageSpeed audit needed for mobile performance baseline."
    ),
    "eeat_links": (
        "No /about.html page exists — critical E-E-A-T gap. "
        "No author bios on blog posts. No testimonials. "
        "0 backlinks tracked. Domain authority unknown."
    ),
    "conversion": (
        "GA4: 95 conversions from ai-publisher-finder.html (top converter). "
        "Homepage: 76 sessions, 0 conversions — needs CTA. "
        "19 pages with 5+ sessions but 0 conversions. "
        "Traffic sources: 52 organic, 113 direct, 12 social, 1 ChatGPT referral."
    ),
}

# ── The 8 compressed mega-days ──────────────────────────────────────

COMPRESSED_DAYS = [
    {
        "day": 8,
        "week": 2,
        "module": "INDEXING EMERGENCY + GA4 CLEANUP",
        "type": "SEO",
        "title": "Fix all indexing blockers + clean GA4 + submit all unindexed pages",
        "detail": (
            "MORNING: (1) Update sitemap.xml — refresh all lastmod dates to today, verify all 46 pages are listed. "
            "(2) Add hreflang tags to all 19 pages missing them + add x-default to ALL 46 pages. "
            "(3) Fix canonical URL inconsistencies across all pages + update copyright 2025→2026 everywhere. "
            "AFTERNOON: (4) Deploy all fixes to Vercel. "
            "(5) Submit ALL unindexed pages via GSC URL Inspection — batch 10 at a time (GSC daily limit ~10-12). "
            "(6) GA4 cleanup: filter bot traffic (77 '(not set)' sessions are bots), set up REAL conversion tracking for contact form + WhatsApp clicks + phone calls. "
            "EVENING: Verify deployment, re-submit sitemap in GSC, document baseline indexing count."
        ),
        "prompt": (
            "Act as a technical SEO specialist. I need to fix indexing for freemindconsult.com. "
            "Currently only 15 of 46 pages have search impressions. "
            "Generate: (1) complete hreflang tag set for 46 pages, (2) canonical tag audit checklist, "
            "(3) GA4 bot filter configuration, (4) conversion tracking setup for contact forms."
        ),
        "why": EVIDENCE["indexing"],
        "est_time": "6 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 8-14. This is the #1 priority — nothing else matters if pages aren't indexed. "
            f"[{datetime.now():%Y-%m-%d}] GSC: 15/31 pages have impressions. GA4: 77 bot sessions to filter."
        ),
        "tool": "VS Code, GSC URL Inspection, GA4 Admin",
    },
    {
        "day": 9,
        "week": 2,
        "module": "TITLES + META DESCRIPTIONS (ALL 46 PAGES)",
        "type": "SEO",
        "title": "Rewrite all titles (<60 chars) + all meta descriptions (<155 chars) + deploy",
        "detail": (
            "MORNING: (1) Audit all 46 titles — currently 43 exceed 60 chars (ALL truncated in Google). "
            "(2) Rewrite homepage title to target 'Free Mind Consultancy' brand + primary service. "
            "(3) Rewrite 5 persona page titles (authors, founders, coaches, professors, govt officers). "
            "(4) Rewrite 8 blog post titles — include '2026', primary keyword near start. "
            "AFTERNOON: (5) Rewrite all 27 service + meme + clinic page titles. "
            "(6) Write meta descriptions for ALL 46 pages — each under 155 chars with CTA. "
            "(7) Update og:title, og:description, twitter:title, twitter:description to match. "
            "EVENING: (8) Deploy all changes. (9) Re-request indexing for top 10 pages."
        ),
        "prompt": (
            "Act as an SEO copywriter. Rewrite title tags and meta descriptions for freemindconsult.com. "
            "Rules: titles under 60 chars, metas under 155 chars with CTA. "
            "Data: 8 pages have CTR < 2% — these need the most compelling titles. "
            "Homepage CTR is 1.4% at position 6.4 (expected: 5%). "
            "Include brand suffix '| Free Mind Consultancy' on service pages."
        ),
        "why": EVIDENCE["titles_meta"],
        "est_time": "5 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 29-35. Moved UP because title/meta fixes are the fastest way to improve CTR. "
            f"[{datetime.now():%Y-%m-%d}] 43/46 titles truncated. 8 pages below 2% CTR. Homepage CTR gap = 3.6%."
        ),
        "tool": "VS Code, GSC",
    },
    {
        "day": 10,
        "week": 2,
        "module": "KEYWORD RESEARCH + QUICK WINS",
        "type": "SEO",
        "title": "Map GSC queries to pages + optimize striking-distance keywords + fix brand terms",
        "detail": (
            "MORNING: (1) Export full GSC query data (82 unique queries). "
            "(2) Map every real query to its landing page — find mismatches. "
            "(3) QUICK WIN: 'isbn registration india' at position 10 with 35 impressions — optimize that page NOW "
            "(add keyword to H1, first paragraph, strengthen content). "
            "(4) QUICK WIN: 'google ads for doctors pune' at position 7 with 8 impressions — optimize clinic page. "
            "AFTERNOON: (5) Generate seed keywords for all verticals (publishing, clinic, meme, AI). "
            "(6) Classify all keywords by intent (informational/commercial/navigational). "
            "(7) Build keyword map: primary + secondary keyword for each of 46 pages. "
            "(8) Address brand confusion: 'freemind chatbot platform' from US generates 200+ impressions, 0 clicks — "
            "add disclaimer or clarify in meta description that we're a consultancy, not a chatbot."
        ),
        "prompt": (
            "Act as an SEO keyword strategist. Analyze these real GSC queries for freemindconsult.com: "
            "82 queries, avg position 35.0. Top opportunities: isbn registration (pos 10, 35 imp), "
            "how to publish a book india (pos 72, 81 imp), google ads for doctors (pos 7, 8 imp). "
            "Create: keyword map for 46 pages, intent classification, content gap analysis."
        ),
        "why": EVIDENCE["keywords"],
        "est_time": "5 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 17-18 + 22-28. Data-first approach — we have real GSC data, no need to guess. "
            f"[{datetime.now():%Y-%m-%d}] 82 queries, avg pos 35.0. isbn at pos 10. Brand confusion from US traffic."
        ),
        "tool": "GSC, SEO Agent data, keyword tools",
    },
    {
        "day": 11,
        "week": 2,
        "module": "CONTENT QUALITY + ON-PAGE + AEO",
        "type": "SEO + AEO",
        "title": "Fix all H1s + add FAQs to all blogs + GEO optimize service pages + fix thin content",
        "detail": (
            "MORNING: (1) Fix hidden H1 on homepage — make it visible and keyword-rich. "
            "(2) Audit heading structure (H1/H2/H3) on all 46 pages — fix hierarchy violations. "
            "(3) Add FAQ sections with FAQPage schema to all 8 blog posts — target People Also Ask. "
            "AFTERNOON: (4) Optimize top 5 service pages for GEO — clear definitions in first 100 words, "
            "structured data-rich content that AI engines can cite. "
            "(5) Add content images to text-heavy service pages (currently pure text). "
            "(6) Expand thin content pages — multiple pages under 800 words need more depth. "
            "(7) Optimize for AI Overviews: add conversational Q&A, structured lists, direct answers."
        ),
        "prompt": (
            "Act as a content SEO specialist. For freemindconsult.com: "
            "(1) Write FAQ sections (5 Q&As each) for 8 blog posts about book publishing, ISBN, ghostwriting, "
            "clinic automation, meme marketing. Target People Also Ask boxes. "
            "(2) Write GEO-optimized opening paragraphs for top 5 service pages."
        ),
        "why": EVIDENCE["content"],
        "est_time": "6 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 36-42 + 85-91. H1 fix + FAQs = highest content ROI. "
            f"[{datetime.now():%Y-%m-%d}] 161 technical issues. Multiple missing H1s and thin content pages."
        ),
        "tool": "VS Code, Schema.org, Rich Results Test",
    },
    {
        "day": 12,
        "week": 2,
        "module": "TECHNICAL SEO + PERFORMANCE + SCHEMA",
        "type": "Technical",
        "title": "PageSpeed audit + image optimization + schema markup + mobile usability + all technical fixes",
        "detail": (
            "MORNING: (1) Run PageSpeed Insights on all priority pages (MOBILE). Record baseline scores. "
            "(2) Audit ALL images — convert to WebP, add width/height attributes, compress to <200KB. "
            "(3) Add missing alt text to all images (detected: multiple images without alt). "
            "(4) Defer non-critical JS + implement font-display:swap for fonts. "
            "AFTERNOON: (5) Audit third-party scripts: Apollo tracker, Meta Pixel, GA4 loading strategy. "
            "(6) Add preconnect hints for Google Fonts, analytics, CDNs. "
            "(7) Add JSON-LD schema to ALL pages missing it — Service schema for service pages, "
            "Article schema for blogs, LocalBusiness for homepage, FAQPage where FAQs exist. "
            "(8) Verify www/non-www redirects + HTTPS enforcement. "
            "EVENING: (9) Re-test PageSpeed. (10) Run Rich Results Test on all pages with schema."
        ),
        "prompt": (
            "Act as a technical SEO engineer. For freemindconsult.com: "
            "(1) Generate JSON-LD schema for: LocalBusiness (homepage), Service (15 service pages), "
            "Article (8 blog posts), FAQPage (blog posts with FAQs). "
            "(2) Create image optimization checklist. (3) JS audit recommendations."
        ),
        "why": EVIDENCE["technical"],
        "est_time": "6 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 43-49 + 68-69 + 120-121. "
            "Schema + speed in one day = compound ranking boost. "
            f"[{datetime.now():%Y-%m-%d}] Missing schema on many pages. Alt text gaps. No PageSpeed baseline."
        ),
        "tool": "PageSpeed Insights, VS Code, Schema.org, WebP converter",
    },
    {
        "day": 13,
        "week": 2,
        "module": "INTERNAL LINKING + TOPIC CLUSTERS + LEAD CAPTURE",
        "type": "SEO + CRO",
        "title": "Build internal link architecture + topic clusters + CTAs + lead capture on all pages",
        "detail": (
            "MORNING: (1) Map current internal link architecture — crawl all 46 pages. "
            "(2) Identify orphan pages (16 pages with 0 search impressions + direct_only pages). "
            "(3) Design 3 topic clusters: Book Publishing pillar (6 cluster pages), "
            "Clinic Automation pillar (8 cluster pages), Meme Marketing pillar (4 cluster pages). "
            "(4) Add 3-5 strategic internal links to each of 8 blog posts. "
            "AFTERNOON: (5) Cross-link service pages ↔ persona pages. Add 'Related Services' sections. "
            "(6) Add clear CTAs to top 5 pages: contact form link, WhatsApp button, phone number. "
            "(7) Add lead capture to AI Publisher Finder (95 conversions — top converter, needs email gate). "
            "(8) Add India-focused local SEO signals: city names, 'India' in content, local schema."
        ),
        "prompt": (
            "Act as a conversion rate optimizer and internal linking specialist. "
            "For freemindconsult.com: (1) Design internal linking strategy connecting 46 pages "
            "into 3 topic clusters. (2) Write CTA copy for WhatsApp button and contact forms. "
            "(3) Design lead capture flow for AI Publisher Finder (95 conversions/month, no email capture)."
        ),
        "why": EVIDENCE["internal_links"] + " | " + EVIDENCE["conversion"],
        "est_time": "6 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 15-16, 19, 50-56, 94-96. "
            "Internal linking + CTAs = fastest path to leads. "
            f"[{datetime.now():%Y-%m-%d}] 16 orphan pages. 19 pages with 0 conversions. "
            "AI Publisher Finder = 95 conversions but no email capture."
        ),
        "tool": "VS Code, Screaming Frog / manual crawl",
    },
    {
        "day": 14,
        "week": 2,
        "module": "E-E-A-T + CONTENT CREATION + LINK BUILDING",
        "type": "SEO + Content",
        "title": "Create About page + author bios + write 3 new blog posts + start link building",
        "detail": (
            "MORNING: (1) Create /about.html — founder bio, team, credentials, Person schema. "
            "(2) Add author bio with credentials to all 8 blog posts + update Article schema with author. "
            "(3) Add visible 'Last Updated: 2026' dates to all blog posts and service pages. "
            "(4) Add testimonials/case study section to homepage and top 3 service pages. "
            "AFTERNOON: (5) Write and publish blog post: 'Meme Marketing for Small Businesses India 2026' (2000 words). "
            "(6) Write and publish blog post: 'Clinic Automation for Doctors India 2026' (2000 words). "
            "(7) Write and publish blog post: 'Business Automation for Indian Startups 2026' (2000 words). "
            "EVENING: (8) Submit 3 new posts to GSC + update sitemap. "
            "(9) Create outreach email templates + send first 5 link building emails. "
            "(10) Submit site to 10 Indian business/startup directories."
        ),
        "prompt": (
            "Act as an SEO content strategist. Write 3 SEO-optimized blog posts (2000 words each) for: "
            "(1) 'Meme Marketing for Small Businesses India 2026' targeting 'meme marketing india' (11 imp, pos 54). "
            "(2) 'Clinic Automation for Doctors India 2026' targeting 'appointment automation clinic' (18 imp, pos 31). "
            "(3) 'Business Automation for Indian Startups 2026' targeting 'business automation services' (2 imp, pos 83). "
            "Include FAQ sections, internal links to service pages, CTAs."
        ),
        "why": EVIDENCE["eeat_links"],
        "est_time": "8 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 57-77 + 64-70 + 92-93 + 113-118. "
            "E-E-A-T + fresh content + first backlinks = authority boost. "
            f"[{datetime.now():%Y-%m-%d}] No About page. No author bios. 0 backlinks. "
            "Content gaps: meme marketing, clinic automation, business automation."
        ),
        "tool": "VS Code, GSC, email, directories",
    },
    {
        "day": 15,
        "week": 2,
        "module": "COMPREHENSIVE REVIEW + PHASE 2 PLANNING",
        "type": "Analytics + Strategy",
        "title": "15-Day performance review vs goals + set Phase 2 strategy + maintenance system",
        "detail": (
            "MORNING: (1) Pull fresh GA4 + GSC data (run SEO Agent pipeline). "
            "(2) Compare: organic sessions before vs after. Indexed pages before vs after. "
            "(3) Check keyword movements — did isbn improve? Did brand confusion reduce? "
            "(4) Measure CTR improvements from title/meta rewrites. "
            "(5) Count leads generated from new CTAs and lead capture. "
            "AFTERNOON: (6) Update keyword map with new GSC data — find new opportunities. "
            "(7) Content audit: which posts drive traffic? Which need improvement? "
            "(8) Set new SMART goals for Phase 2 (Days 16-45). "
            "(9) Plan content calendar for Phase 2: 8-12 new blog post topics from content gaps. "
            "(10) Build Phase 2 link building + technical maintenance schedule. "
            "EVENING: (11) Run SEO Agent Intelligence Engine to generate updated roadmap. "
            "(12) Set up ongoing SEO maintenance: monthly data pull, quarterly audit, weekly GSC check."
        ),
        "prompt": (
            "Act as an SEO Director. Review 15 days of SEO work on freemindconsult.com. "
            "Baseline: 52 organic sessions, 15 indexed pages, avg position 35. "
            "Evaluate: indexing progress, CTR improvements, content impact, lead generation. "
            "Create Phase 2 plan: next 30 days of highest-impact actions."
        ),
        "why": (
            "15-day sprint review ensures we measure impact before planning next phase. "
            "Evidence-based planning prevents wasted effort. "
            "The SEO Agent Intelligence Engine automates ongoing analysis."
        ),
        "est_time": "4 hours",
        "fixed_flexible": "Fixed",
        "status": "",
        "notes": (
            "COMBINES original Days 78-84 + 99-105 + 124-126. "
            "Sprint review → set new targets → automate with SEO Agent. "
            f"[{datetime.now():%Y-%m-%d}] Baseline: 52 organic sessions, 333 total sessions, "
            "4 total clicks from search, 497 impressions, 15/31 pages visible in search."
        ),
        "tool": "SEO Agent, GA4, GSC, Excel",
    },
]

# ── Compressed Weekly Themes ────────────────────────────────────────

COMPRESSED_THEMES = [
    {
        "week": 1,
        "days": "1-7",
        "phase": "Foundations + Critical Fixes [DONE]",
        "focus": "SEO mindset, fix broken canonicals, fix missing tags, set SMART goals, verify GSC setup",
        "deliverables": "3 critical bugs fixed, SMART goals set, GSC baseline documented",
        "mix": "SEO 100%",
    },
    {
        "week": 2,
        "days": "8-15",
        "phase": "FULL IMPLEMENTATION SPRINT",
        "focus": (
            "Indexing emergency, title/meta rewrite, keyword research, content quality, "
            "technical SEO, internal linking, E-E-A-T, content creation, link building, "
            "lead capture, performance review"
        ),
        "deliverables": (
            "All pages indexed, 46 titles/metas rewritten, keyword map complete, "
            "FAQs on all blogs, schema on all pages, 3 new blog posts, "
            "About page live, CTAs on all pages, link building started, "
            "15-day review + Phase 2 plan"
        ),
        "mix": "SEO 50% Technical 20% Content 15% CRO 10% Analytics 5%",
    },
]


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler()],
    )
    logger = logging.getLogger("compress")

    # 1. Backup
    timestamp = datetime.now().strftime("%Y_%m_%d_%H%M%S")
    backup_path = BACKUPS_DIR / f"SEO_Roadmap_PRE_COMPRESS_{timestamp}.xlsx"
    shutil.copy2(WORKBOOK, backup_path)
    logger.info("Backup: %s", backup_path.name)

    # 2. Load workbook (preserving everything)
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["101-Day Roadmap"]

    # 3. Capture formatting from an existing data row (row 2 = Day 1)
    ref_fonts = {}
    ref_aligns = {}
    for col in range(1, 15):
        cell = ws.cell(row=2, column=col)
        ref_fonts[col] = cell.font.copy()
        ref_aligns[col] = cell.alignment.copy() if cell.alignment else Alignment()

    # 4. Delete old pending rows (rows 9 through max_row) — Days 8-126
    # We delete from bottom up to avoid row index shifting
    for row_idx in range(ws.max_row, 8, -1):
        ws.delete_rows(row_idx, 1)
    logger.info("Cleared old pending tasks (rows 9+)")

    # 5. Write the 8 compressed mega-days (rows 9-16)
    for i, day_data in enumerate(COMPRESSED_DAYS):
        row_idx = 9 + i
        values = [
            day_data["day"],
            "",
            day_data["week"],
            day_data["module"],
            day_data["type"],
            day_data["title"],
            day_data["detail"],
            day_data["prompt"],
            day_data["why"],
            day_data["est_time"],
            day_data["fixed_flexible"],
            day_data["status"],
            day_data["notes"],
            day_data["tool"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = ref_fonts.get(col_idx, Font())
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    logger.info("Wrote 8 compressed mega-days (rows 9-16)")

    # 6. Update Weekly Themes sheet
    ws_themes = wb["Weekly Themes"]
    # Clear existing data rows (keep header at row 1)
    for row_idx in range(ws_themes.max_row, 1, -1):
        ws_themes.delete_rows(row_idx, 1)

    for i, theme in enumerate(COMPRESSED_THEMES):
        row_idx = 2 + i
        theme_vals = [
            theme["week"],
            theme["days"],
            theme["phase"],
            theme["focus"],
            theme["deliverables"],
            theme["mix"],
        ]
        for col_idx, val in enumerate(theme_vals, start=1):
            cell = ws_themes.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    logger.info("Updated Weekly Themes sheet (2 weeks)")

    # 7. Save
    wb.save(WORKBOOK)
    wb.close()
    logger.info("Saved compressed roadmap: %s", WORKBOOK.name)
    logger.info("")
    logger.info("DONE: 126 days → 15 days (7 DONE + 8 mega-days)")
    logger.info("Backup at: %s", backup_path)


if __name__ == "__main__":
    main()
